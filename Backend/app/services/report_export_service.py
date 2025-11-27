"""
Report Export Service - Generate PDF and Excel reports with charts
"""
import io
import os
from datetime import datetime
from typing import List, Optional
import matplotlib
matplotlib.use('Agg')  # Non-GUI backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.traffic_record import TrafficStatistics, HourlyStatistics, DailyTrend, RoadComparison


class ReportExportService:
    """Service for exporting traffic reports to PDF and Excel with visualizations"""

    @staticmethod
    def generate_pdf_report(
        statistics: List[TrafficStatistics],
        hourly_trends: List[HourlyStatistics],
        daily_trends: List[DailyTrend],
        road_comparisons: List[RoadComparison],
        start_date: str,
        end_date: str
    ) -> bytes:
        """
        Generate comprehensive PDF report with charts

        Returns:
            PDF file as bytes
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=0.5*inch, rightMargin=0.5*inch)
        story = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1E40AF'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#3B82F6'),
            spaceAfter=12,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )

        # Title
        title = Paragraph("BAO CAO GIAO THONG THONG MINH", title_style)
        story.append(title)

        # Report info
        info_text = f"""
        <b>Thoi gian:</b> {start_date} den {end_date}<br/>
        <b>Ngay tao bao cao:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>
        <b>So luong duong giam sat:</b> {len(statistics)}<br/>
        """
        story.append(Paragraph(info_text, styles['Normal']))
        story.append(Spacer(1, 0.3 * inch))

        # Statistics Summary Table
        story.append(Paragraph("1. Tong Quan Thong Ke", heading_style))

        if statistics:
            # Prepare table data
            table_data = [[
                'Ten Duong',
                'So Ban Ghi',
                'TB So Xe',
                'Max Xe',
                'TB Toc Do',
                'Gio Cao Diem',
                'Ty Le Tac (%)'
            ]]

            for stat in statistics:
                table_data.append([
                    stat.road_name,
                    str(stat.total_records),
                    f"{stat.avg_vehicles:.1f}",
                    str(stat.max_vehicles),
                    f"{stat.avg_speed:.1f} km/h",
                    f"{stat.peak_hour}:00" if stat.peak_hour is not None else "N/A",
                    f"{stat.congestion_rate:.1f}%"
                ])

            # Create table with better column widths
            col_widths = [1.5*inch, 0.8*inch, 0.7*inch, 0.6*inch, 0.9*inch, 1*inch, 0.9*inch]
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.3 * inch))

        # Track temporary files to delete after PDF is built
        temp_files = []

        # Generate and add hourly chart
        if hourly_trends and len(hourly_trends) > 0:
            story.append(Paragraph("2. Xu Huong Theo Gio", heading_style))
            hourly_chart_path = ReportExportService._create_hourly_chart(hourly_trends)
            if hourly_chart_path and os.path.exists(hourly_chart_path):
                try:
                    img = Image(hourly_chart_path, width=6*inch, height=3.5*inch)
                    story.append(img)
                    story.append(Spacer(1, 0.2 * inch))
                    temp_files.append(hourly_chart_path)
                except Exception as e:
                    print(f"Error adding hourly chart to PDF: {e}")

        # Generate and add daily trend chart
        if daily_trends and len(daily_trends) > 0:
            story.append(Paragraph("3. Xu Huong Theo Ngay", heading_style))
            daily_chart_path = ReportExportService._create_daily_chart(daily_trends)
            if daily_chart_path and os.path.exists(daily_chart_path):
                try:
                    img = Image(daily_chart_path, width=6*inch, height=3.5*inch)
                    story.append(img)
                    story.append(Spacer(1, 0.2 * inch))
                    temp_files.append(daily_chart_path)
                except Exception as e:
                    print(f"Error adding daily chart to PDF: {e}")

        # Road comparison chart
        if road_comparisons and len(road_comparisons) > 1:
            story.append(PageBreak())
            story.append(Paragraph("4. So Sanh Cac Tuyen Duong", heading_style))
            comparison_chart_path = ReportExportService._create_comparison_chart(road_comparisons)
            if comparison_chart_path and os.path.exists(comparison_chart_path):
                try:
                    img = Image(comparison_chart_path, width=6*inch, height=3.5*inch)
                    story.append(img)
                    temp_files.append(comparison_chart_path)
                except Exception as e:
                    print(f"Error adding comparison chart to PDF: {e}")

        # Build PDF
        try:
            doc.build(story)
            buffer.seek(0)
            pdf_data = buffer.getvalue()

            # Clean up temporary files after PDF is built
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except Exception as e:
                    print(f"Warning: Could not delete temp file {temp_file}: {e}")

            return pdf_data
        except Exception as e:
            print(f"Error building PDF: {e}")
            # Clean up temp files even if there's an error
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                except:
                    pass
            raise

    @staticmethod
    def _create_hourly_chart(hourly_trends: List[HourlyStatistics]) -> Optional[str]:
        """Create hourly trend chart"""
        try:
            plt.clf()
            plt.close('all')

            fig, ax1 = plt.subplots(figsize=(10, 6))

            hours = [h.hour for h in hourly_trends]
            vehicles = [h.avg_vehicles for h in hourly_trends]
            speeds = [h.avg_speed for h in hourly_trends]

            # Plot vehicles
            color = 'tab:blue'
            ax1.set_xlabel('Gio trong ngay', fontsize=12)
            ax1.set_ylabel('So xe trung binh', color=color, fontsize=12)
            ax1.bar(hours, vehicles, color=color, alpha=0.6, label='So xe')
            ax1.tick_params(axis='y', labelcolor=color)
            ax1.set_xticks(range(0, 24, 2))

            # Plot speeds on second y-axis
            ax2 = ax1.twinx()
            color = 'tab:red'
            ax2.set_ylabel('Toc do TB (km/h)', color=color, fontsize=12)
            ax2.plot(hours, speeds, color=color, marker='o', linewidth=2, label='Toc do')
            ax2.tick_params(axis='y', labelcolor=color)

            plt.title('Luu Luong va Toc Do Theo Gio', fontsize=14, fontweight='bold')
            fig.tight_layout()

            # Save to temp file with absolute path
            import tempfile
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, f'temp_hourly_{datetime.now().timestamp()}.png')
            plt.savefig(temp_file, dpi=150, bbox_inches='tight')
            plt.close(fig)

            return temp_file

        except Exception as e:
            print(f"Error creating hourly chart: {e}")
            plt.close('all')
            return None

    @staticmethod
    def _create_daily_chart(daily_trends: List[DailyTrend]) -> Optional[str]:
        """Create daily trend chart"""
        try:
            plt.clf()
            plt.close('all')

            fig, ax = plt.subplots(figsize=(10, 6))

            dates = [datetime.strptime(d.date, '%Y-%m-%d') for d in daily_trends]
            vehicles = [d.avg_vehicles for d in daily_trends]
            max_vehicles = [d.max_vehicles for d in daily_trends]

            ax.plot(dates, vehicles, marker='o', linewidth=2, label='TB so xe', color='tab:blue')
            ax.plot(dates, max_vehicles, marker='s', linewidth=2, label='Max so xe', color='tab:orange')

            ax.set_xlabel('Ngay', fontsize=12)
            ax.set_ylabel('So xe', fontsize=12)
            ax.set_title('Xu Huong Luu Luong Theo Ngay', fontsize=14, fontweight='bold')
            ax.legend()
            ax.grid(True, alpha=0.3)

            # Format x-axis
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
            plt.xticks(rotation=45)

            fig.tight_layout()

            # Save to temp file with absolute path
            import tempfile
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, f'temp_daily_{datetime.now().timestamp()}.png')
            plt.savefig(temp_file, dpi=150, bbox_inches='tight')
            plt.close(fig)

            return temp_file

        except Exception as e:
            print(f"Error creating daily chart: {e}")
            plt.close('all')
            return None

    @staticmethod
    def _create_comparison_chart(comparisons: List[RoadComparison]) -> Optional[str]:
        """Create road comparison chart"""
        try:
            plt.clf()
            plt.close('all')

            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5))

            roads = [c.road_name for c in comparisons]
            vehicles = [c.avg_vehicles for c in comparisons]
            congestion = [c.congestion_rate for c in comparisons]

            # Bar chart for vehicles
            ax1.bar(roads, vehicles, color='tab:blue', alpha=0.7)
            ax1.set_xlabel('Tuyen duong', fontsize=11)
            ax1.set_ylabel('So xe trung binh', fontsize=11)
            ax1.set_title('So Sanh Luu Luong', fontsize=12, fontweight='bold')
            plt.setp(ax1.xaxis.get_majorticklabels(), rotation=45, ha='right')

            # Bar chart for congestion
            colors_list = ['green' if c < 30 else 'orange' if c < 60 else 'red' for c in congestion]
            ax2.bar(roads, congestion, color=colors_list, alpha=0.7)
            ax2.set_xlabel('Tuyen duong', fontsize=11)
            ax2.set_ylabel('Ty le tac nghen (%)', fontsize=11)
            ax2.set_title('So Sanh Tac Nghen', fontsize=12, fontweight='bold')
            plt.setp(ax2.xaxis.get_majorticklabels(), rotation=45, ha='right')

            fig.tight_layout()

            # Save to temp file with absolute path
            import tempfile
            temp_dir = tempfile.gettempdir()
            temp_file = os.path.join(temp_dir, f'temp_comparison_{datetime.now().timestamp()}.png')
            plt.savefig(temp_file, dpi=150, bbox_inches='tight')
            plt.close(fig)

            return temp_file

        except Exception as e:
            print(f"Error creating comparison chart: {e}")
            plt.close('all')
            return None

    @staticmethod
    def generate_excel_report(
        statistics: List[TrafficStatistics],
        hourly_trends: List[HourlyStatistics],
        daily_trends: List[DailyTrend],
        road_comparisons: List[RoadComparison],
        start_date: str,
        end_date: str
    ) -> bytes:
        """
        Generate comprehensive Excel report with multiple sheets and charts

        Returns:
            Excel file as bytes
        """
        buffer = io.BytesIO()
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Sheet 1: Summary
        ws_summary = wb.create_sheet("Tong Quan")
        ReportExportService._create_summary_sheet(ws_summary, statistics, start_date, end_date)

        # Sheet 2: Hourly Trends
        if hourly_trends and len(hourly_trends) > 0:
            ws_hourly = wb.create_sheet("Xu Huong Theo Gio")
            ReportExportService._create_hourly_sheet(ws_hourly, hourly_trends)

        # Sheet 3: Daily Trends
        if daily_trends and len(daily_trends) > 0:
            ws_daily = wb.create_sheet("Xu Huong Theo Ngay")
            ReportExportService._create_daily_sheet(ws_daily, daily_trends)

        # Sheet 4: Road Comparison
        if road_comparisons and len(road_comparisons) > 0:
            ws_comparison = wb.create_sheet("So Sanh Tuyen Duong")
            ReportExportService._create_comparison_sheet(ws_comparison, road_comparisons)

        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def _create_summary_sheet(ws, statistics: List[TrafficStatistics], start_date: str, end_date: str):
        """Create summary sheet in Excel"""
        # Title
        ws['A1'] = 'BAO CAO GIAO THONG THONG MINH'
        ws['A1'].font = Font(size=18, bold=True, color='1E40AF')
        ws.merge_cells('A1:G1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Info
        ws['A3'] = f'Thoi gian: {start_date} den {end_date}'
        ws['A4'] = f'Ngay tao: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'

        # Headers
        headers = ['Ten Duong', 'So Ban Ghi', 'TB So Xe', 'Max Xe', 'TB Toc Do (km/h)', 'Gio Cao Diem', 'Ty Le Tac (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data
        for row, stat in enumerate(statistics, 7):
            ws.cell(row=row, column=1, value=stat.road_name)
            ws.cell(row=row, column=2, value=stat.total_records)
            ws.cell(row=row, column=3, value=round(stat.avg_vehicles, 1))
            ws.cell(row=row, column=4, value=stat.max_vehicles)
            ws.cell(row=row, column=5, value=round(stat.avg_speed, 1))
            ws.cell(row=row, column=6, value=f"{stat.peak_hour}:00" if stat.peak_hour is not None else "N/A")
            ws.cell(row=row, column=7, value=round(stat.congestion_rate, 1))

        # Auto-adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18

    @staticmethod
    def _create_hourly_sheet(ws, hourly_trends: List[HourlyStatistics]):
        """Create hourly trends sheet with chart"""
        # Headers
        headers = ['Gio', 'TB So Xe', 'TB Toc Do (km/h)', 'Trang Thai']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data
        for row, trend in enumerate(hourly_trends, 2):
            ws.cell(row=row, column=1, value=f"{trend.hour}:00")
            ws.cell(row=row, column=2, value=round(trend.avg_vehicles, 1))
            ws.cell(row=row, column=3, value=round(trend.avg_speed, 1))
            ws.cell(row=row, column=4, value=trend.traffic_status)

        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Create line chart
        chart = LineChart()
        chart.title = "Luu Luong Theo Gio"
        chart.y_axis.title = "So xe"
        chart.x_axis.title = "Gio"
        chart.height = 10
        chart.width = 20

        data = Reference(ws, min_col=2, min_row=1, max_row=len(hourly_trends) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(hourly_trends) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "F2")

    @staticmethod
    def _create_daily_sheet(ws, daily_trends: List[DailyTrend]):
        """Create daily trends sheet with chart"""
        # Headers
        headers = ['Ngay', 'TB So Xe', 'Max So Xe', 'TB Toc Do (km/h)', 'Gio Cao Diem']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data
        for row, trend in enumerate(daily_trends, 2):
            ws.cell(row=row, column=1, value=trend.date)
            ws.cell(row=row, column=2, value=round(trend.avg_vehicles, 1))
            ws.cell(row=row, column=3, value=trend.max_vehicles)
            ws.cell(row=row, column=4, value=round(trend.avg_speed, 1))
            ws.cell(row=row, column=5, value=f"{trend.peak_hour}:00")

        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Create line chart
        chart = LineChart()
        chart.title = "Xu Huong Theo Ngay"
        chart.y_axis.title = "So xe"
        chart.x_axis.title = "Ngay"
        chart.height = 10
        chart.width = 20

        data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=len(daily_trends) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(daily_trends) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "G2")

    @staticmethod
    def _create_comparison_sheet(ws, comparisons: List[RoadComparison]):
        """Create road comparison sheet with charts"""
        # Headers
        headers = ['Ten Duong', 'TB So Xe', 'TB Toc Do (km/h)', 'Ty Le Tac (%)', 'So Ban Ghi']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

        # Data
        for row, comp in enumerate(comparisons, 2):
            ws.cell(row=row, column=1, value=comp.road_name)
            ws.cell(row=row, column=2, value=round(comp.avg_vehicles, 1))
            ws.cell(row=row, column=3, value=round(comp.avg_speed, 1))
            ws.cell(row=row, column=4, value=round(comp.congestion_rate, 1))
            ws.cell(row=row, column=5, value=comp.total_records)

        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

        # Create bar chart
        chart = BarChart()
        chart.title = "So Sanh Luu Luong"
        chart.y_axis.title = "So xe"
        chart.x_axis.title = "Tuyen duong"
        chart.height = 10
        chart.width = 20

        data = Reference(ws, min_col=2, min_row=1, max_row=len(comparisons) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(comparisons) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "G2")


# Global service instance
report_export_service = ReportExportService()
